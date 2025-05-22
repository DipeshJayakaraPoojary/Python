import pandas as pd
from openpyxl import load_workbook
import os

# === CONFIGURATION ===
data_file = "data.xlsx"         # Your input data file
excel_file = "output.xlsx"     # Your existing Excel workbook
sheet_name = "Sheet1"          # Target sheet

# === Step 1: Read and clean the data ===
df = pd.read_csv(data_file, sep='\t')  # Tab-separated

# Replace 'NULL' strings with actual NaNs
df.replace("NULL", pd.NA, inplace=True)

# Convert columns to numeric
numeric_cols = ['lotsize', 'tradedlots', 'numberoftrades', 'tradedvolume']
for col in numeric_cols:
    if col in df.columns:
        df[col] = pd.to_numeric(df[col], errors='coerce')

# Convert date columns
date_cols = ['publicationdate', 'deliverystart', 'deliveryend']
for col in date_cols:
    if col in df.columns:
        df[col] = pd.to_datetime(df[col], format='%d-%m-%Y', errors='coerce')

# === Step 2: Append to existing Excel sheet ===
if not os.path.exists(excel_file):
    # If Excel file doesn't exist, create it
    df.to_excel(excel_file, sheet_name=sheet_name, index=False)
    print("Excel file created.")
else:
    # Load existing workbook
    book = load_workbook(excel_file)
    writer = pd.ExcelWriter(excel_file, engine='openpyxl', mode='a', if_sheet_exists='overlay')
    writer.book = book

    # Determine start row to append after existing data
    if sheet_name in book.sheetnames:
        start_row = book[sheet_name].max_row
    else:
        start_row = 0

    # Append data
    df.to_excel(writer, sheet_name=sheet_name, index=False, header=False, startrow=start_row)
    writer.close()
    print("Data appended to Excel.")
